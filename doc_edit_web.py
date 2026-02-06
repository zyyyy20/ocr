import argparse
import base64
import datetime as _dt
import io
import json
import mimetypes
import os
import re
import threading
import traceback
import uuid
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from urllib.parse import parse_qs, urlparse

import openpyxl
from openpyxl.utils import range_boundaries


_RE_NUMERIC_HINT = re.compile(r"(分|成绩|score|平时|总分|得分)", re.IGNORECASE)
_RE_INT_HINT = re.compile(r"(班级|学号|编号|序号)", re.IGNORECASE)

def _set_paddle_env_flags():
    os.environ.setdefault("FLAGS_use_mkldnn", "0")
    os.environ.setdefault("FLAGS_use_onednn", "0")
    os.environ.setdefault("FLAGS_enable_onednn", "0")
    os.environ.setdefault("FLAGS_disable_mkldnn", "1")
    os.environ.setdefault("FLAGS_enable_pir_in_executor", "0")
    os.environ.setdefault("FLAGS_use_new_executor", "0")


def _now_iso():
    return _dt.datetime.now().isoformat(timespec="seconds")


def _json_bytes(obj):
    return json.dumps(obj, ensure_ascii=False).encode("utf-8")


def _read_body(handler):
    length = int(handler.headers.get("Content-Length", "0") or "0")
    return handler.rfile.read(length) if length > 0 else b""


def _parse_json(handler):
    body = _read_body(handler)
    if not body:
        return None
    return json.loads(body.decode("utf-8"))

def _uploads_dir():
    p = os.path.join(os.path.dirname(__file__), "uploads")
    os.makedirs(p, exist_ok=True)
    return p


def _decode_data_url_or_b64(data_url):
    s = (data_url or "").strip()
    if not s:
        raise ValueError("empty upload")
    mime = None
    if s.startswith("data:"):
        head, _, tail = s.partition(",")
        if ";base64" not in head:
            raise ValueError("only base64 data url supported")
        mime = head[5:].split(";")[0].strip() or None
        raw = base64.b64decode(tail.encode("utf-8"), validate=False)
        return raw, mime
    raw = base64.b64decode(s.encode("utf-8"), validate=False)
    return raw, mime


def _write_bytes(path, data):
    with open(path, "wb") as f:
        f.write(data)


def _convert_pdf_to_png_bytes(pdf_bytes):
    try:
        import fitz

        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        page = doc.load_page(0)
        pix = page.get_pixmap(dpi=200)
        return pix.tobytes("png")
    except Exception:
        pass

    try:
        from pdf2image import convert_from_bytes

        images = convert_from_bytes(pdf_bytes, dpi=200)
        if not images:
            raise RuntimeError("empty pdf")
        out = io.BytesIO()
        images[0].save(out, format="PNG")
        return out.getvalue()
    except Exception:
        raise RuntimeError("PDF 转 PNG 需要安装 PyMuPDF(fitz) 或 pdf2image")


def _convert_svg_to_png_bytes(svg_bytes):
    try:
        import cairosvg

        return cairosvg.svg2png(bytestring=svg_bytes)
    except Exception:
        raise RuntimeError("SVG 转 PNG 需要安装 cairosvg")


def _safe_int(v, default=None):
    try:
        return int(v)
    except Exception:
        return default


def _coerce_value(v):
    if v is None:
        return ""
    if isinstance(v, (int, float)):
        return v
    return str(v)


def _infer_table_from_sheet(ws):
    dim = ws.calculate_dimension()
    if dim == "A1":
        if ws["A1"].value is None:
            return None
    min_col, min_row, max_col, max_row = range_boundaries(dim)

    header_row = None
    for r in range(min_row, max_row + 1):
        values = [ws.cell(row=r, column=c).value for c in range(min_col, max_col + 1)]
        if any(v is not None and str(v).strip() != "" for v in values):
            header_row = r
            break
    if header_row is None:
        return None

    last_row = header_row
    for r in range(header_row + 1, max_row + 1):
        values = [ws.cell(row=r, column=c).value for c in range(min_col, max_col + 1)]
        if any(v is not None and str(v).strip() != "" for v in values):
            last_row = r
        else:
            break

    return {
        "min_row": header_row,
        "min_col": min_col,
        "max_col": max_col,
        "max_row": last_row,
    }


def _normalize_headers(values):
    headers = []
    seen = {}
    for idx, v in enumerate(values):
        name = str(v).strip() if v is not None and str(v).strip() != "" else f"列{idx + 1}"
        base = name
        if base in seen:
            seen[base] += 1
            name = f"{base}_{seen[base]}"
        else:
            seen[base] = 1
        headers.append(name)
    return headers


def _parse_a1_cells(text):
    text = (text or "").strip()
    if not text:
        return []
    if ":" in text:
        min_col, min_row, max_col, max_row = range_boundaries(text)
        out = []
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                out.append(openpyxl.utils.get_column_letter(c) + str(r))
        return out
    cells = []
    for part in re.split(r"[,\s]+", text):
        part = part.strip()
        if part:
            cells.append(part)
    return cells


def _extract_excel(input_path, sheet_name=None, table_range=None, text_cells=None):
    wb = openpyxl.load_workbook(input_path)
    ws = wb[sheet_name] if sheet_name else wb.active

    if table_range:
        min_col, min_row, max_col, max_row = range_boundaries(table_range)
        table = {"min_row": min_row, "min_col": min_col, "max_row": max_row, "max_col": max_col}
    else:
        table = _infer_table_from_sheet(ws)
        if table is None:
            raise RuntimeError("未在工作表中检测到表格区域，请通过 --table-range 指定。")

    header_row = table["min_row"]
    headers_raw = [ws.cell(row=header_row, column=c).value for c in range(table["min_col"], table["max_col"] + 1)]
    headers = _normalize_headers(headers_raw)

    data_rows = []
    for r in range(header_row + 1, table["max_row"] + 1):
        row = {}
        empty = True
        for i, c in enumerate(range(table["min_col"], table["max_col"] + 1)):
            v = ws.cell(row=r, column=c).value
            if v is not None and str(v).strip() != "":
                empty = False
            row[headers[i]] = _coerce_value(v)
        if empty:
            continue
        data_rows.append(row)

    cells = _parse_a1_cells(text_cells)
    text_lines = []
    for addr in cells:
        try:
            v = ws[addr].value
        except Exception:
            v = None
        if v is None:
            text_lines.append("")
        else:
            text_lines.append(str(v))
    text = "\n".join(text_lines).rstrip("\n")

    meta = {
        "type": "excel",
        "input_path": input_path,
        "sheet": ws.title,
        "table": {
            "header_row": header_row,
            "min_col": table["min_col"],
            "max_col": table["max_col"],
            "data_start_row": header_row + 1,
            "data_end_row": table["max_row"],
        },
        "text_cells": cells,
    }
    return headers, data_rows, text, meta


def _cluster_centers(values, tol):
    values = sorted(values)
    clusters = []
    for v in values:
        if not clusters:
            clusters.append([v])
            continue
        if abs(v - (sum(clusters[-1]) / len(clusters[-1]))) <= tol:
            clusters[-1].append(v)
        else:
            clusters.append([v])
    centers = [sum(c) / len(c) for c in clusters]
    return centers


def _assign_to_nearest(centers, v):
    best = None
    best_i = None
    for i, c in enumerate(centers):
        d = abs(v - c)
        if best is None or d < best:
            best = d
            best_i = i
    return best_i


def _extract_image_table(input_path):
    _set_paddle_env_flags()
    from paddleocr import PaddleOCR

    ocr = PaddleOCR(lang="ch")
    result = ocr.predict(input_path, use_textline_orientation=True)
    item = result[0] if result else None
    if item is None or not hasattr(item, "get"):
        raise RuntimeError("未获取到 OCR 结果。")
    boxes = item.get("rec_boxes") or item.get("rec_polys") or item.get("dt_polys") or []
    texts = item.get("rec_texts") or []
    scores = item.get("rec_scores") or []
    if not texts:
        raise RuntimeError("未识别到文本。")

    rows_y = []
    cols_x = []
    items = []
    for i in range(min(len(texts), len(boxes))):
        b = boxes[i]
        try:
            pts = list(b)
        except Exception:
            continue
        if len(pts) == 4 and hasattr(pts[0], "__len__") and len(pts[0]) == 2:
            xs = [float(p[0]) for p in pts]
            ys = [float(p[1]) for p in pts]
        elif len(pts) == 4 and all(isinstance(v, (int, float)) for v in pts):
            x1, y1, x2, y2 = [float(v) for v in pts]
            xs = [x1, x2, x2, x1]
            ys = [y1, y1, y2, y2]
        elif len(pts) == 8 and all(isinstance(v, (int, float)) for v in pts):
            xs = [float(pts[0]), float(pts[2]), float(pts[4]), float(pts[6])]
            ys = [float(pts[1]), float(pts[3]), float(pts[5]), float(pts[7])]
        else:
            continue

        xc = sum(xs) / len(xs)
        yc = sum(ys) / len(ys)
        rows_y.append(yc)
        cols_x.append(xc)
        items.append({"xc": xc, "yc": yc, "text": str(texts[i]), "score": float(scores[i]) if i < len(scores) else None})

    if not items:
        raise RuntimeError("OCR 结果结构无法解析。")

    y_tol = max(8.0, (max(rows_y) - min(rows_y)) / 80.0)
    x_tol = max(8.0, (max(cols_x) - min(cols_x)) / 40.0)
    row_centers = _cluster_centers(rows_y, y_tol)
    col_centers = _cluster_centers(cols_x, x_tol)

    grid = [[[] for _ in range(len(col_centers))] for __ in range(len(row_centers))]
    for it in items:
        ri = _assign_to_nearest(row_centers, it["yc"])
        ci = _assign_to_nearest(col_centers, it["xc"])
        if ri is None or ci is None:
            continue
        grid[ri][ci].append(it)

    table_rows = []
    for r in range(len(row_centers)):
        row = []
        for c in range(len(col_centers)):
            cell_items = sorted(grid[r][c], key=lambda x: x["xc"])
            row.append(" ".join(i["text"] for i in cell_items).strip())
        if any(v.strip() for v in row):
            table_rows.append(row)

    if not table_rows:
        raise RuntimeError("未能重建表格。")

    headers = _normalize_headers(table_rows[0])
    data = []
    for r in table_rows[1:]:
        r2 = (r + [""] * len(headers))[: len(headers)]
        data.append({headers[i]: r2[i] for i in range(len(headers))})

    meta = {
        "type": "image",
        "input_path": input_path,
        "sheet": None,
        "table": None,
        "text_cells": [],
    }
    text = "\n".join(["\t".join(row) for row in table_rows]).strip()
    return headers, data, text, meta


def _build_rules(headers, rows):
    sample = rows[:50]
    rules = []
    for h in headers:
        values = [r.get(h) for r in sample]
        non_empty = [v for v in values if v is not None and str(v).strip() != ""]
        numeric_like = False
        int_like = False
        if _RE_NUMERIC_HINT.search(h):
            numeric_like = True
        if _RE_INT_HINT.search(h):
            int_like = True

        if non_empty:
            ok_num = 0
            ok_int = 0
            for v in non_empty:
                try:
                    fv = float(v)
                    ok_num += 1
                    if abs(fv - int(fv)) < 1e-9:
                        ok_int += 1
                except Exception:
                    pass
            if ok_num == len(non_empty):
                numeric_like = True
            if ok_int == len(non_empty):
                int_like = True

        required = h in ("姓名", "班级") or bool(_RE_INT_HINT.search(h))
        if h.startswith("备注"):
            required = False

        rule = {"name": h, "required": bool(required), "type": "string"}
        if int_like:
            rule["type"] = "int"
        elif numeric_like:
            rule["type"] = "number"

        if _RE_NUMERIC_HINT.search(h):
            rule["min"] = 0
            rule["max"] = 100
        rules.append(rule)
    return rules


def _validate(headers, rows, rules):
    errors = []
    name_to_rule = {r["name"]: r for r in rules}
    for idx, row in enumerate(rows):
        for h in headers:
            rule = name_to_rule.get(h) or {"name": h, "required": False, "type": "string"}
            v = row.get(h, "")
            s = "" if v is None else str(v).strip()
            if rule.get("required") and s == "":
                errors.append({"row": idx + 1, "col": h, "msg": "必填"})
                continue
            if s == "":
                continue
            t = rule.get("type")
            if t == "int":
                iv = _safe_int(s, default=None)
                if iv is None:
                    errors.append({"row": idx + 1, "col": h, "msg": "需为整数"})
                    continue
                vnum = float(iv)
            elif t == "number":
                try:
                    vnum = float(s)
                except Exception:
                    errors.append({"row": idx + 1, "col": h, "msg": "需为数字"})
                    continue
            else:
                vnum = None

            if vnum is not None:
                mn = rule.get("min")
                mx = rule.get("max")
                if mn is not None and vnum < float(mn):
                    errors.append({"row": idx + 1, "col": h, "msg": f"需 ≥ {mn}"})
                if mx is not None and vnum > float(mx):
                    errors.append({"row": idx + 1, "col": h, "msg": f"需 ≤ {mx}"})
    return errors


def _copy_cell_style(src_cell, dst_cell):
    dst_cell._style = src_cell._style
    dst_cell.number_format = src_cell.number_format
    dst_cell.font = src_cell.font
    dst_cell.fill = src_cell.fill
    dst_cell.border = src_cell.border
    dst_cell.alignment = src_cell.alignment
    dst_cell.protection = src_cell.protection
    dst_cell.comment = src_cell.comment


def _export_excel_from_excel_source(meta, headers, rows, text, output_name):
    input_path = meta["input_path"]
    wb = openpyxl.load_workbook(input_path)
    ws = wb[meta["sheet"]]

    t = meta["table"]
    header_row = int(t["header_row"])
    min_col = int(t["min_col"])
    max_col = int(t["max_col"])
    data_start = int(t["data_start_row"])
    data_end = int(t["data_end_row"])
    current_capacity = max(0, data_end - data_start + 1)

    row_count = len(rows)
    need_capacity = max(current_capacity, row_count)
    template_row = data_start if current_capacity == 0 else data_start

    if need_capacity > current_capacity:
        insert_at = data_end + 1
        insert_n = need_capacity - current_capacity
        if insert_n > 0:
            ws.insert_rows(insert_at, amount=insert_n)
            for r in range(insert_at, insert_at + insert_n):
                for c in range(min_col, max_col + 1):
                    _copy_cell_style(ws.cell(row=template_row, column=c), ws.cell(row=r, column=c))

    name_to_col = {}
    headers_raw = [ws.cell(row=header_row, column=c).value for c in range(min_col, max_col + 1)]
    normalized = _normalize_headers(headers_raw)
    for i, h in enumerate(normalized):
        name_to_col[h] = min_col + i

    for i in range(need_capacity):
        excel_row = data_start + i
        if i < row_count:
            record = rows[i]
        else:
            record = {}
        for h in headers:
            c = name_to_col.get(h)
            if c is None:
                continue
            v = record.get(h, "")
            cell = ws.cell(row=excel_row, column=c)
            if v is None or (isinstance(v, str) and v.strip() == ""):
                cell.value = None
            else:
                if isinstance(v, (int, float)):
                    cell.value = v
                else:
                    s = str(v).strip()
                    if s == "":
                        cell.value = None
                    else:
                        iv = _safe_int(s, default=None)
                        if iv is not None and str(iv) == s:
                            cell.value = iv
                        else:
                            try:
                                fv = float(s)
                                cell.value = fv
                            except Exception:
                                cell.value = s

    cells = meta.get("text_cells") or []
    lines = (text or "").splitlines()
    for i, addr in enumerate(cells):
        v = lines[i] if i < len(lines) else ""
        try:
            ws[addr].value = v if v.strip() != "" else None
        except Exception:
            pass

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue(), output_name


def _export_excel_from_image_source(headers, rows, text, output_name):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "OCR"

    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i).value = h

    for r, record in enumerate(rows, start=2):
        for i, h in enumerate(headers, start=1):
            ws.cell(row=r, column=i).value = record.get(h, "")

    if text and text.strip():
        start_row = len(rows) + 4
        ws.cell(row=start_row - 1, column=1).value = "文本内容"
        for i, line in enumerate(text.splitlines(), start=0):
            ws.cell(row=start_row + i, column=1).value = line

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue(), output_name


def _html_page():
    return """<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>文档编辑器</title>
  <link rel="stylesheet" href="https://unpkg.com/tabulator-tables@6.2.5/dist/css/tabulator.min.css">
  <style>
    body{font-family:system-ui,-apple-system,Segoe UI,Roboto,Helvetica,Arial,sans-serif;margin:0;padding:0;background:#fafafa}
    header{padding:12px 16px;background:#111827;color:#fff}
    main{padding:16px;max-width:1400px;margin:0 auto}
    .row{display:flex;gap:12px;flex-wrap:wrap;align-items:flex-start}
    .card{background:#fff;border:1px solid #e5e7eb;border-radius:8px;padding:12px;flex:1;min-width:320px}
    .toolbar{display:flex;gap:8px;flex-wrap:wrap;margin-bottom:10px}
    button{padding:8px 10px;border:1px solid #d1d5db;background:#fff;border-radius:6px;cursor:pointer}
    button.primary{background:#2563eb;border-color:#2563eb;color:#fff}
    button.danger{background:#dc2626;border-color:#dc2626;color:#fff}
    button:disabled{opacity:.5;cursor:not-allowed}
    textarea{width:100%;min-height:220px;resize:vertical;border:1px solid #d1d5db;border-radius:6px;padding:8px}
    .meta{color:#6b7280;font-size:12px}
    .errors{color:#b91c1c;white-space:pre-wrap;font-size:12px}
    .history{font-size:12px;max-height:220px;overflow:auto;border:1px solid #e5e7eb;border-radius:6px;padding:8px;background:#f9fafb}
    .footer{margin-top:12px;color:#6b7280;font-size:12px}
    .drop{border:1px dashed #9ca3af;border-radius:8px;padding:12px;background:#f9fafb}
    .drop.drag{background:#eef2ff;border-color:#6366f1}
    .preview{width:100%;height:520px;border:1px solid #e5e7eb;border-radius:8px;overflow:hidden;background:#fff}
    .preview img{width:100%;height:100%;object-fit:contain;display:block}
    .preview object{width:100%;height:100%;display:block}
    input[type=text]{padding:7px 8px;border:1px solid #d1d5db;border-radius:6px}
  </style>
</head>
<body>
  <header>
    <div>文档编辑器（导入 → OCR/解析 → 交互编辑 → 导出 Excel）</div>
    <div class="meta" id="meta">未加载文件</div>
  </header>
  <main>
    <div class="row">
      <div class="card" style="flex: 1;min-width:360px;max-width:520px">
        <div class="toolbar">
          <div class="meta">原始文件导入（PNG/SVG/PDF/XLSX）</div>
        </div>
        <div id="drop" class="drop">
          <div class="toolbar" style="margin-bottom:8px">
            <input id="fileInput" type="file" accept=".png,.jpg,.jpeg,.bmp,.tif,.tiff,.svg,.pdf,.xlsx,.xlsm" />
            <button id="uploadBtn" class="primary">导入并解析</button>
          </div>
          <div class="toolbar" style="margin-bottom:8px">
            <input id="sheet" type="text" placeholder="sheet（Excel 可选）" />
            <input id="tableRange" type="text" placeholder="table-range（如 A1:H200）" />
            <input id="textCells" type="text" placeholder="text-cells（如 J1:J10）" />
          </div>
          <div class="meta" id="uploadStatus">选择文件后点击“导入并解析”</div>
        </div>
        <div style="height:10px"></div>
        <div class="toolbar">
          <div class="meta">原始文件预览</div>
        </div>
        <div class="preview" id="preview"></div>
      </div>
      <div class="card" style="flex: 2;min-width:520px">
        <div class="toolbar">
          <button id="addRow">添加行</button>
          <button id="deleteRow" class="danger">删除选中行</button>
          <button id="undo">撤销</button>
          <button id="redo">重做</button>
          <button id="export" class="primary">导出 Excel</button>
        </div>
        <div id="table"></div>
        <div class="row" style="margin-top:12px">
          <div class="card" style="flex: 1;min-width:360px">
            <div class="toolbar">
              <div class="meta">文本内容（可编辑）</div>
            </div>
            <textarea id="textContent" placeholder="在这里编辑文本内容"></textarea>
          </div>
          <div class="card" style="flex: 1;min-width:360px">
            <div class="toolbar">
              <div class="meta">数据校验</div>
            </div>
            <div id="errors" class="errors"></div>
            <div class="toolbar" style="margin-top:10px">
              <div class="meta">修改历史</div>
            </div>
            <div id="history" class="history"></div>
          </div>
        </div>
        <div class="footer">双击单元格可编辑；点击行可选中后删除；每列表头支持过滤；导出前会做校验。</div>
      </div>
    </div>
  </main>
  <script src="https://unpkg.com/tabulator-tables@6.2.5/dist/js/tabulator.min.js"></script>
  <script src="/app.js"></script>
</body>
</html>"""


def _js_app():
    return r"""let state = null;
let table = null;
let undoStack = [];
let redoStack = [];
let history = [];
let rules = [];
let uiBound = false;

function deepCopy(x){ return JSON.parse(JSON.stringify(x)); }
function nowISO(){ return new Date().toISOString().replace(/\.\d+Z$/,''); }

function setMeta(text){ document.getElementById("meta").textContent = text || ""; }
function setStatus(text){ document.getElementById("uploadStatus").textContent = text || ""; }

function recordHistory(entry){
  history.push(entry);
  if(history.length > 5000) history.shift();
  renderHistory();
}

function renderHistory(){
  const el = document.getElementById("history");
  const lines = history.slice(-200).map(h => `[${h.ts}] ${h.action} ${h.detail}`);
  el.textContent = lines.join("\n");
}

function updateButtons(){
  document.getElementById("undo").disabled = undoStack.length === 0;
  document.getElementById("redo").disabled = redoStack.length === 0;
  const hasTable = !!table;
  document.getElementById("addRow").disabled = !hasTable;
  document.getElementById("deleteRow").disabled = !hasTable;
  document.getElementById("export").disabled = !hasTable;
}

function pushSnapshot(reason){
  if(!state) return;
  undoStack.push({ts: nowISO(), reason, snapshot: deepCopy({rows: state.rows, text: state.text})});
  if(undoStack.length > 200) undoStack.shift();
  redoStack = [];
  updateButtons();
}

function validateClient(){
  if(!state) return [];
  const errs = [];
  const byName = {};
  for(const r of rules) byName[r.name] = r;
  for(let i=0;i<state.rows.length;i++){
    const row = state.rows[i];
    for(const h of state.headers){
      const rule = byName[h] || {name:h, required:false, type:"string"};
      const v = row[h];
      const s = (v === null || v === undefined) ? "" : String(v).trim();
      if(rule.required && s === ""){
        errs.push({row:i+1, col:h, msg:"必填"});
        continue;
      }
      if(s === "") continue;
      let num = null;
      if(rule.type === "int"){
        if(!/^-?\d+$/.test(s)){ errs.push({row:i+1, col:h, msg:"需为整数"}); continue; }
        num = Number.parseInt(s,10);
      }else if(rule.type === "number"){
        if(!/^-?\d+(\.\d+)?$/.test(s)){ errs.push({row:i+1, col:h, msg:"需为数字"}); continue; }
        num = Number.parseFloat(s);
      }
      if(num !== null){
        if(rule.min !== undefined && num < Number(rule.min)) errs.push({row:i+1, col:h, msg:`需 ≥ ${rule.min}`});
        if(rule.max !== undefined && num > Number(rule.max)) errs.push({row:i+1, col:h, msg:`需 ≤ ${rule.max}`});
      }
    }
  }
  const el = document.getElementById("errors");
  if(errs.length === 0){
    el.textContent = "";
  }else{
    el.textContent = errs.slice(0,200).map(e => `第${e.row}行 ${e.col}: ${e.msg}`).join("\n");
  }
  return errs;
}

function destroyTable(){
  if(table){
    try{ table.destroy(); }catch(e){}
    table = null;
  }
}

function renderLocalPreview(file){
  const preview = document.getElementById("preview");
  preview.innerHTML = "";
  if(!file) return;
  const url = URL.createObjectURL(file);
  const name = (file.name || "").toLowerCase();
  if(name.endsWith(".pdf") || file.type === "application/pdf"){
    const obj = document.createElement("object");
    obj.data = url;
    obj.type = "application/pdf";
    preview.appendChild(obj);
  }else if(name.endsWith(".svg") || file.type.includes("svg")){
    const obj = document.createElement("object");
    obj.data = url;
    obj.type = "image/svg+xml";
    preview.appendChild(obj);
  }else{
    if(file.type && file.type.startsWith("image/")){
      const img = document.createElement("img");
      img.src = url;
      preview.appendChild(img);
    }else{
      const div = document.createElement("div");
      div.className = "meta";
      div.style.padding = "10px";
      div.textContent = "该文件类型无法在浏览器直接预览。";
      preview.appendChild(div);
    }
  }
}

function renderServerPreview(){
  const preview = document.getElementById("preview");
  preview.innerHTML = "";
  if(!state || !state.preview) return;
  const kind = state.preview.original_kind || "original";
  const mime = state.preview.original_mime || "";
  const url = state.preview.original_url || "/api/file?kind=" + encodeURIComponent(kind);
  if(mime.includes("pdf")){
    const obj = document.createElement("object");
    obj.data = url;
    obj.type = mime;
    preview.appendChild(obj);
  }else if(mime.includes("svg")){
    const obj = document.createElement("object");
    obj.data = url;
    obj.type = mime;
    preview.appendChild(obj);
  }else{
    if(mime.startsWith("image/")){
      const img = document.createElement("img");
      img.src = url;
      preview.appendChild(img);
    }else{
      const div = document.createElement("div");
      div.className = "meta";
      div.style.padding = "10px";
      div.textContent = "该文件类型无法在浏览器直接预览。";
      preview.appendChild(div);
    }
  }
}

function initEditor(payload){
  destroyTable();
  state = payload;
  rules = state.rules || [];
  undoStack = [];
  redoStack = [];
  history = [];
  renderHistory();

  setMeta(`${state.meta.type} | ${state.meta.input_name} | sheet: ${state.meta.sheet || '-'} | ${state.headers.length} 列 | ${state.rows.length} 行`);
  document.getElementById("textContent").value = state.text || "";

  const cols = state.headers.map(h => ({title: h, field: h, editor: "input", headerFilter: "input"}));
  table = new Tabulator("#table", {
    height: "650px",
    data: state.rows,
    layout: "fitDataStretch",
    columns: cols,
    selectable: 1,
    clipboard: true,
    clipboardCopyHeader: true,
    cellEdited: function(cell){
      const field = cell.getField();
      const row = cell.getRow().getPosition(true);
      recordHistory({ts: nowISO(), action:"edit", detail:`row ${row+1} col ${field}`});
      state.rows = table.getData();
      validateClient();
    }
  });
  state.rows = table.getData();
  renderServerPreview();
  validateClient();
  updateButtons();
}

async function apiLoad(){
  const res = await fetch("/api/load");
  if(res.status === 404) return null;
  if(!res.ok) throw new Error(await res.text());
  return await res.json();
}

async function uploadFile(file){
  const opts = {
    sheet: document.getElementById("sheet").value || "",
    table_range: document.getElementById("tableRange").value || "",
    text_cells: document.getElementById("textCells").value || "",
  };
  const dataUrl = await new Promise((resolve, reject) => {
    const reader = new FileReader();
    reader.onload = () => resolve(String(reader.result || ""));
    reader.onerror = () => reject(reader.error || new Error("read failed"));
    reader.readAsDataURL(file);
  });

  const payload = {filename: file.name || "upload", data_url: dataUrl, options: opts};
  const resp = await fetch("/api/upload", {method:"POST", headers:{"Content-Type":"application/json"}, body: JSON.stringify(payload)});
  if(!resp.ok) throw new Error(await resp.text());
  return await resp.json();
}

function bindUI(){
  if(uiBound) return;
  uiBound = true;

  const fileInput = document.getElementById("fileInput");
  const uploadBtn = document.getElementById("uploadBtn");
  const drop = document.getElementById("drop");

  fileInput.addEventListener("change", () => {
    const f = fileInput.files && fileInput.files[0];
    renderLocalPreview(f);
  });

  uploadBtn.addEventListener("click", async () => {
    const f = fileInput.files && fileInput.files[0];
    if(!f){
      alert("请先选择文件");
      return;
    }
    try{
      setStatus("正在解析，请稍候...");
      uploadBtn.disabled = true;
      const payload = await uploadFile(f);
      initEditor(payload);
      setStatus("解析完成");
      recordHistory({ts: nowISO(), action:"load", detail: payload.meta.input_name});
    }catch(e){
      setStatus("解析失败");
      document.getElementById("errors").textContent = String(e && e.stack ? e.stack : e);
    }finally{
      uploadBtn.disabled = false;
    }
  });

  drop.addEventListener("dragenter", (e) => { e.preventDefault(); drop.classList.add("drag"); });
  drop.addEventListener("dragover", (e) => { e.preventDefault(); drop.classList.add("drag"); });
  drop.addEventListener("dragleave", (e) => { e.preventDefault(); drop.classList.remove("drag"); });
  drop.addEventListener("drop", (e) => {
    e.preventDefault();
    drop.classList.remove("drag");
    const f = e.dataTransfer && e.dataTransfer.files && e.dataTransfer.files[0];
    if(f){
      fileInput.files = e.dataTransfer.files;
      renderLocalPreview(f);
    }
  });

  document.getElementById("textContent").addEventListener("input", () => {
    if(!state) return;
    state.text = document.getElementById("textContent").value;
    recordHistory({ts: nowISO(), action:"text", detail:"edit text"});
  });

  document.getElementById("addRow").addEventListener("click", () => {
    if(!table || !state) return;
    pushSnapshot("addRow");
    const empty = {};
    for(const h of state.headers) empty[h] = "";
    table.addRow(empty, true);
    state.rows = table.getData();
    recordHistory({ts: nowISO(), action:"add", detail:"add row"});
    validateClient();
  });

  document.getElementById("deleteRow").addEventListener("click", () => {
    if(!table || !state) return;
    const selected = table.getSelectedRows();
    if(selected.length === 0) return;
    pushSnapshot("deleteRow");
    const idx = selected[0].getPosition(true);
    selected[0].delete();
    state.rows = table.getData();
    recordHistory({ts: nowISO(), action:"delete", detail:`delete row ${idx+1}`});
    validateClient();
  });

  document.getElementById("undo").addEventListener("click", () => {
    if(!table || !state) return;
    if(undoStack.length === 0) return;
    const cur = deepCopy({rows: state.rows, text: state.text});
    const last = undoStack.pop();
    redoStack.push({ts: nowISO(), reason: last.reason, snapshot: cur});
    state.rows = deepCopy(last.snapshot.rows);
    state.text = last.snapshot.text;
    table.replaceData(state.rows);
    document.getElementById("textContent").value = state.text || "";
    recordHistory({ts: nowISO(), action:"undo", detail:last.reason});
    validateClient();
    updateButtons();
  });

  document.getElementById("redo").addEventListener("click", () => {
    if(!table || !state) return;
    if(redoStack.length === 0) return;
    const cur = deepCopy({rows: state.rows, text: state.text});
    const last = redoStack.pop();
    undoStack.push({ts: nowISO(), reason: last.reason, snapshot: cur});
    state.rows = deepCopy(last.snapshot.rows);
    state.text = last.snapshot.text;
    table.replaceData(state.rows);
    document.getElementById("textContent").value = state.text || "";
    recordHistory({ts: nowISO(), action:"redo", detail:last.reason});
    validateClient();
    updateButtons();
  });

  document.getElementById("export").addEventListener("click", async () => {
    if(!table || !state) return;
    state.rows = table.getData();
    state.text = document.getElementById("textContent").value;
    const errs = validateClient();
    if(errs.length > 0){
      alert("存在校验错误，修复后再导出。");
      return;
    }
    const payload = {headers: state.headers, rows: state.rows, text: state.text, history};
    const resp = await fetch("/api/export", {method:"POST", headers:{"Content-Type":"application/json"}, body: JSON.stringify(payload)});
    if(!resp.ok){
      const msg = await resp.text();
      alert("导出失败：" + msg);
      return;
    }
    const blob = await resp.blob();
    const cd = resp.headers.get("Content-Disposition") || "";
    let filename = "export.xlsx";
    const m = cd.match(/filename\*=UTF-8''([^;]+)/) || cd.match(/filename=\"?([^\";]+)\"?/);
    if(m) filename = decodeURIComponent(m[1]);
    const url = URL.createObjectURL(blob);
    const a = document.createElement("a");
    a.href = url;
    a.download = filename;
    document.body.appendChild(a);
    a.click();
    a.remove();
    URL.revokeObjectURL(url);
    recordHistory({ts: nowISO(), action:"export", detail:filename});
  });
}

(async () => {
  bindUI();
  updateButtons();
  try{
    const payload = await apiLoad();
    if(payload){
      initEditor(payload);
      setStatus("已加载服务器当前文件");
    }else{
      setStatus("选择文件后点击“导入并解析”");
    }
  }catch(e){
    document.getElementById("errors").textContent = String(e && e.stack ? e.stack : e);
  }
})();"""


class _App:
    def __init__(self, input_path=None, sheet=None, table_range=None, text_cells=None):
        self._lock = threading.Lock()
        self.headers = []
        self.rows = []
        self.text = ""
        self.meta = None
        self.rules = []
        self.preview = None
        self._original_file = None
        self._ocr_image_file = None

        if input_path:
            self.load_from_path(input_path, sheet=sheet, table_range=table_range, text_cells=text_cells)

    def _set_doc(self, headers, rows, text, meta, original_file, ocr_image_file):
        self.headers = headers
        self.rows = rows
        self.text = text or ""
        self.meta = meta
        self.rules = _build_rules(headers, rows)
        self._original_file = original_file
        self._ocr_image_file = ocr_image_file

        self.preview = None
        if original_file:
            self.preview = {
                "original_kind": "original",
                "original_url": "/api/file?kind=original",
                "original_mime": original_file.get("mime") or "application/octet-stream",
            }
        if ocr_image_file:
            self.preview["ocr_kind"] = "ocr"
            self.preview["ocr_url"] = "/api/file?kind=ocr"
            self.preview["ocr_mime"] = ocr_image_file.get("mime") or "image/png"

    def load_from_path(self, input_path, sheet=None, table_range=None, text_cells=""):
        with self._lock:
            ext = os.path.splitext(input_path)[1].lower()
            name = os.path.basename(input_path)

            if ext in (".xlsx", ".xlsm"):
                headers, rows, text, meta = _extract_excel(input_path, sheet, table_range, text_cells)
                meta["input_name"] = name
                self._set_doc(headers, rows, text, meta, original_file=None, ocr_image_file=None)
                return self.load_payload()

            if ext in (".png", ".jpg", ".jpeg", ".bmp", ".tif", ".tiff"):
                headers, rows, text, meta = _extract_image_table(input_path)
                meta["input_name"] = name
                self._set_doc(
                    headers,
                    rows,
                    text,
                    meta,
                    original_file={"path": input_path, "mime": mimetypes.guess_type(name)[0]},
                    ocr_image_file={"path": input_path, "mime": mimetypes.guess_type(name)[0]},
                )
                return self.load_payload()

            raise RuntimeError("仅支持 xlsx/xlsm 或 png/jpg/bmp/tif。")

    def load_from_upload(self, filename, data_url, options):
        raw, mime = _decode_data_url_or_b64(data_url)
        name = os.path.basename(filename or "upload")
        ext = os.path.splitext(name)[1].lower()
        if not ext and mime:
            ext = mimetypes.guess_extension(mime) or ""

        doc_id = uuid.uuid4().hex
        up_dir = _uploads_dir()
        original_path = os.path.join(up_dir, f"{doc_id}_original{ext or ''}")
        _write_bytes(original_path, raw)

        original_file = {"path": original_path, "mime": mime or mimetypes.guess_type(name)[0]}
        ocr_image_file = None

        with self._lock:
            if ext in (".xlsx", ".xlsm"):
                sheet = (options or {}).get("sheet") or None
                table_range = (options or {}).get("table_range") or None
                text_cells = (options or {}).get("text_cells") or ""
                headers, rows, text, meta = _extract_excel(original_path, sheet, table_range, text_cells)
                meta["input_name"] = name
                self._set_doc(headers, rows, text, meta, original_file=original_file, ocr_image_file=None)
                return self.load_payload()

            if ext in (".png", ".jpg", ".jpeg", ".bmp", ".tif", ".tiff"):
                ocr_image_file = original_file
                headers, rows, text, meta = _extract_image_table(original_path)
                meta["input_name"] = name
                self._set_doc(headers, rows, text, meta, original_file=original_file, ocr_image_file=ocr_image_file)
                return self.load_payload()

            if ext == ".pdf":
                png_bytes = _convert_pdf_to_png_bytes(raw)
                ocr_path = os.path.join(up_dir, f"{doc_id}_ocr.png")
                _write_bytes(ocr_path, png_bytes)
                ocr_image_file = {"path": ocr_path, "mime": "image/png"}
                headers, rows, text, meta = _extract_image_table(ocr_path)
                meta["input_name"] = name
                meta["source_ext"] = ext
                self._set_doc(headers, rows, text, meta, original_file=original_file, ocr_image_file=ocr_image_file)
                return self.load_payload()

            if ext == ".svg":
                png_bytes = _convert_svg_to_png_bytes(raw)
                ocr_path = os.path.join(up_dir, f"{doc_id}_ocr.png")
                _write_bytes(ocr_path, png_bytes)
                ocr_image_file = {"path": ocr_path, "mime": "image/png"}
                headers, rows, text, meta = _extract_image_table(ocr_path)
                meta["input_name"] = name
                meta["source_ext"] = ext
                self._set_doc(headers, rows, text, meta, original_file=original_file, ocr_image_file=ocr_image_file)
                return self.load_payload()

            raise RuntimeError("不支持的文件类型")

    def load_payload(self):
        with self._lock:
            if not self.meta:
                raise FileNotFoundError("no document loaded")
            return {
                "headers": self.headers,
                "rows": self.rows,
                "text": self.text,
                "meta": self.meta,
                "rules": self.rules,
                "preview": self.preview,
            }

    def get_file(self, kind):
        with self._lock:
            if kind == "ocr":
                f = self._ocr_image_file
            else:
                f = self._original_file
            if not f or not f.get("path") or not os.path.exists(f["path"]):
                raise FileNotFoundError("file missing")
            path = f["path"]
            mime = f.get("mime") or "application/octet-stream"
            with open(path, "rb") as fp:
                return fp.read(), mime

    def export(self, headers, rows, text, history):
        with self._lock:
            if not self.meta:
                raise ValueError("未加载文件")

            errors = _validate(headers, rows, self.rules)
            if errors:
                raise ValueError("存在校验错误")

            ts = _dt.datetime.now().strftime("%Y%m%d_%H%M%S")
            base = os.path.splitext(os.path.basename(self.meta.get("input_name") or self.meta.get("input_path") or "export"))[0]
            output_name = f"{base}_edited_{ts}.xlsx"

            if self.meta["type"] == "excel":
                content, name = _export_excel_from_excel_source(self.meta, headers, rows, text, output_name)
            else:
                content, name = _export_excel_from_image_source(headers, rows, text, output_name)

            hist_path = os.path.join(os.path.dirname(__file__), f"{base}_history_{ts}.json")
            try:
                with open(hist_path, "wb") as f:
                    f.write(_json_bytes({"exported_at": _now_iso(), "history": history}))
            except Exception:
                pass

            return content, name


def _make_handler(app: _App):
    class Handler(BaseHTTPRequestHandler):
        def log_message(self, fmt, *args):
            return

        def _send(self, status, body, content_type="text/plain; charset=utf-8", extra_headers=None):
            self.send_response(status)
            self.send_header("Content-Type", content_type)
            self.send_header("Cache-Control", "no-store")
            if extra_headers:
                for k, v in extra_headers.items():
                    self.send_header(k, v)
            self.end_headers()
            if body:
                self.wfile.write(body)

        def do_GET(self):
            u = urlparse(self.path)
            path = u.path
            if path == "/" or path == "/index.html":
                self._send(HTTPStatus.OK, _html_page().encode("utf-8"), "text/html; charset=utf-8")
                return
            if path == "/app.js":
                self._send(HTTPStatus.OK, _js_app().encode("utf-8"), "application/javascript; charset=utf-8")
                return
            if path == "/api/load":
                try:
                    payload = app.load_payload()
                    self._send(HTTPStatus.OK, _json_bytes(payload), "application/json; charset=utf-8")
                except FileNotFoundError:
                    self._send(HTTPStatus.NOT_FOUND, b"no document loaded")
                except Exception:
                    self._send(HTTPStatus.INTERNAL_SERVER_ERROR, traceback.format_exc().encode("utf-8"))
                return
            if path == "/api/file":
                try:
                    qs = parse_qs(u.query or "")
                    kind = (qs.get("kind") or ["original"])[0]
                    data, mime = app.get_file(kind=kind)
                    self._send(HTTPStatus.OK, data, mime or "application/octet-stream")
                except FileNotFoundError:
                    self._send(HTTPStatus.NOT_FOUND, b"file not found")
                except Exception:
                    self._send(HTTPStatus.INTERNAL_SERVER_ERROR, traceback.format_exc().encode("utf-8"))
                return
            self._send(HTTPStatus.NOT_FOUND, b"not found")

        def do_POST(self):
            path = urlparse(self.path).path
            if path == "/api/upload":
                try:
                    payload = _parse_json(self) or {}
                    filename = payload.get("filename") or "upload"
                    data_url = payload.get("data_url") or ""
                    options = payload.get("options") or {}
                    doc = app.load_from_upload(filename, data_url, options)
                    self._send(HTTPStatus.OK, _json_bytes(doc), "application/json; charset=utf-8")
                except ValueError as e:
                    self._send(HTTPStatus.BAD_REQUEST, str(e).encode("utf-8"))
                except Exception:
                    self._send(HTTPStatus.INTERNAL_SERVER_ERROR, traceback.format_exc().encode("utf-8"))
                return
            if path == "/api/export":
                try:
                    payload = _parse_json(self) or {}
                    headers = payload.get("headers") or []
                    rows = payload.get("rows") or []
                    text = payload.get("text") or ""
                    history = payload.get("history") or []
                    content, name = app.export(headers, rows, text, history)
                    self._send(
                        HTTPStatus.OK,
                        content,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        extra_headers={
                            "Content-Disposition": "attachment; filename*=UTF-8''" + _quote(name),
                        },
                    )
                except ValueError as e:
                    self._send(HTTPStatus.BAD_REQUEST, str(e).encode("utf-8"))
                except Exception:
                    self._send(HTTPStatus.INTERNAL_SERVER_ERROR, traceback.format_exc().encode("utf-8"))
                return
            self._send(HTTPStatus.NOT_FOUND, b"not found")

    return Handler


def _quote(s):
    return re.sub(r"[^A-Za-z0-9._-]", lambda m: "%%%02X" % ord(m.group(0)), s)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", default=None, help="可选：启动时加载的输入文件路径（xlsx/xlsm 或 png/jpg/pdf/svg）")
    parser.add_argument("--sheet", default=None, help="Excel 工作表名称（默认 active）")
    parser.add_argument("--table-range", default=None, help="Excel 表格范围（例如 A1:H20），不填则自动探测")
    parser.add_argument("--text-cells", default="", help="文本单元格地址：如 J1:J10 或 J1,J2,J3")
    parser.add_argument("--host", default="127.0.0.1")
    parser.add_argument("--port", type=int, default=8000)
    args = parser.parse_args()

    _set_paddle_env_flags()
    app = _App(args.input, args.sheet, args.table_range, args.text_cells)
    server = ThreadingHTTPServer((args.host, args.port), _make_handler(app))
    url = f"http://{args.host}:{args.port}/"
    print("server:", url)
    server.serve_forever()


if __name__ == "__main__":
    main()
